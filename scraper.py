import requests
from bs4 import BeautifulSoup
import xlwt
import openpyxl
from openpyxl import load_workbook

# On first run, adds headers
def initWorkbook(book):
    # Initiates Recipe sheet
    RecipeBookSheet = book.create_sheet(title="Recipe")
    RecipeBookSheet.cell(row=1, column=1).value = 'Recipe_ID'
    RecipeBookSheet.cell(row=1, column=2).value = 'Title'
    RecipeBookSheet.cell(row=1, column=3).value = 'PrepTime'
    RecipeBookSheet.cell(row=1, column=4).value = 'CookTime'
    RecipeBookSheet.cell(row=1, column=5).value = 'Difficulty'
    RecipeBookSheet.cell(row=1, column=6).value = 'ServingSize'
    RecipeBookSheet.cell(row=1, column=7).value = 'Description'

    # Initiates Ingredient sheet
    RecipeBookSheet = book.create_sheet(title='Ingredients')
    RecipeBookSheet.cell(row=1, column=1).value = 'Recipe_ID'
    RecipeBookSheet.cell(row=1, column=2).value = 'IngredientName'

    # Initiates Method sheet
    RecipeBookSheet = book.create_sheet(title='Method')
    RecipeBookSheet.cell(row=1, column=1).value = 'Recipe_ID'
    RecipeBookSheet.cell(row=1, column=2).value = 'MethodStep'


ua = {"User-Agent":"Mozilla/5.0"}
url = input ("Enter URL: ")

r = requests.get(url, headers=ua)

soup = BeautifulSoup(r.content, 'html.parser')
recipeid = 11111

try:
    RecipeBook = load_workbook(filename= 'test.xlsx')
    RecipeBookSheet = RecipeBook['Recipe']
    recipeid = RecipeBookSheet.cell(row=RecipeBookSheet.max_row, column=1).value + 1
except:
    RecipeBook = openpyxl.Workbook()
    initWorkbook(RecipeBook)

# Sets numbering style to number
#number = xlwt.XFStyle()
#number.num_format_str = '0'

# Creates workbook and initiates
#RecipeBook = xlwt.Workbook()
#initWorkbook(RecipeBook)

# Assigns active sheet to Recipe
RecipeBookSheet = RecipeBook['Recipe']

# Grabs all the relevant soup elements
title = soup.select_one('.recipe-header__title').text.strip()
prepTime = soup.select_one('.recipe-details__cooking-time-prep > span').text.strip()
cookTime = soup.select_one('.recipe-details__cooking-time-cook > span').text.strip()
difficulty = soup.select_one('.recipe-details__item--skill-level > span').text.strip()
servingsize = soup.select_one('.recipe-details__item--servings > span').text.strip()
description = soup.select_one('.recipe-header__description').text.strip()

# Gets next empty row
index = RecipeBookSheet.max_row + 1

# Writes those soup elements to the sheet
RecipeBookSheet.cell(row=index, column=1).value = recipeid
RecipeBookSheet.cell(row=index, column=2).value = title
RecipeBookSheet.cell(row=index, column=3).value = prepTime
RecipeBookSheet.cell(row=index, column=4).value = cookTime
RecipeBookSheet.cell(row=index, column=5).value = difficulty
RecipeBookSheet.cell(row=index, column=6).value = servingsize
RecipeBookSheet.cell(row=index, column=7).value = description

# Gets the ingredients sheet and relevant soup elements
RecipeBookSheet = RecipeBook['Ingredients']
ingredients = soup.select('.ingredients-list__item')

# Writes the ingredients to the sheet
index = RecipeBookSheet.max_row + 1
for x in ingredients:
    if(len(x.findChildren())==0):
        RecipeBookSheet.cell(row=index, column=1).value = recipeid
        RecipeBookSheet.cell(row=index, column=2).value = x.text.strip()
    else:
        RecipeBookSheet.cell(row=index, column=1).value = recipeid
        RecipeBookSheet.cell(row=index, column=2).value = x.findAll(text=True)[0]+x.findAll(text=True)[1]
    index+=1

# Gets the Method sheet and relevant soup elements
RecipeBookSheet = RecipeBook['Method']
method = soup.select('.method__item p')

# Write sthe method steps to the sheet
index = RecipeBookSheet.max_row + 1
for x in method:
    RecipeBookSheet.cell(row=index, column=1).value = recipeid
    RecipeBookSheet.cell(row=index, column=2).value = x.text.strip()
    index+=1


# Save the Workbook
RecipeBook.save(filename = 'test.xlsx')